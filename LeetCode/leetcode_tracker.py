import pandas as pd
import requests
import re
import os
import time
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill


def get_leetcode_problems():
    """Fetch full list of LeetCode problems with IDs and titles"""
    print("Fetching complete LeetCode problem list...")

    problems_url = "https://leetcode.com/api/problems/all/"

    try:
        response = requests.get(problems_url)

        if response.status_code != 200:
            print(f"Failed to fetch problems: HTTP {response.status_code}")
            return {}

        data = response.json()

        # Create a mapping of problem titles to their IDs and details
        problem_map = {}
        for problem in data.get("stat_status_pairs", []):
            stat = problem.get("stat", {})
            title = stat.get("question__title", "")
            frontend_id = stat.get("frontend_question_id", "")
            difficulty = problem.get("difficulty", {}).get("level", 0)

            # Convert difficulty level to text
            difficulty_text = "Unknown"
            if difficulty == 1:
                difficulty_text = "Easy"
            elif difficulty == 2:
                difficulty_text = "Medium"
            elif difficulty == 3:
                difficulty_text = "Hard"

            # Store in map with title as key
            problem_map[title] = {"id": str(frontend_id), "difficulty": difficulty_text}

            # Also store with ID as key for exact matching
            problem_map[str(frontend_id)] = {
                "title": title,
                "difficulty": difficulty_text,
            }

        print(f"Retrieved details for {len(problem_map) // 2} LeetCode problems")
        return problem_map

    except Exception as e:
        print(f"Error fetching problems: {e}")
        return {}


def extract_problem_titles_from_excel(excel_file):
    """Extract problem titles from the Excel file with links"""
    print(f"Reading problem titles from: {excel_file}")

    try:
        # Load Excel file
        df = pd.read_excel(excel_file, header=None)

        # The first column contains titles with links
        titles_column = df.iloc[:, 0]

        # Extract titles and problem IDs
        problem_titles = []
        for cell in titles_column:
            if pd.notna(cell):
                # Try to extract problem ID or title from the cell
                # Convert to string if it's not already
                cell_str = str(cell)

                # Try to extract problem ID from link if it exists
                id_match = re.search(r"problems/([^/]+)/", cell_str)
                title_match = re.search(r"<(.+?)>", cell_str) or re.search(
                    r"([\w\s\(\)-]+)", cell_str
                )

                problem_info = {}

                # If the cell has a URL, extract slug
                if id_match:
                    problem_info["slug"] = id_match.group(1)

                # Extract title
                if title_match:
                    problem_info["title"] = title_match.group(1).strip()
                else:
                    problem_info["title"] = cell_str.strip()

                problem_titles.append(problem_info)

        print(f"Extracted {len(problem_titles)} problem titles from your Excel file")
        return problem_titles

    except Exception as e:
        print(f"Error extracting titles: {e}")
        return []


def get_user_solved_problems(username):
    """Get the list of problems solved by the user"""
    print(f"Fetching solved problems for: {username}")

    api_url = f"https://leetcode.com/api/problems/all/"
    profile_url = f"https://leetcode.com/{username}"

    solved_problems = []

    try:
        # First try official API
        response = requests.get(api_url, params={"username": username})

        if response.status_code == 200:
            data = response.json()

            # Check if we can find data for the user
            if "user_name" in data and data["user_name"] == username:
                # Extract solved problems
                for problem in data.get("stat_status_pairs", []):
                    if problem.get("status") == "ac":  # ac = Accepted/Solved
                        stat = problem.get("stat", {})
                        solved_problems.append(
                            {
                                "id": str(stat.get("frontend_question_id", "")),
                                "title": stat.get("question__title", ""),
                            }
                        )

                print(f"Found {len(solved_problems)} solved problems via API")
                return solved_problems

        # If API approach fails, try scraping the profile page
        print("API approach failed, trying to scrape profile page...")

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }

        profile_response = requests.get(profile_url, headers=headers)

        if profile_response.status_code == 200:
            soup = BeautifulSoup(profile_response.text, "html.parser")

            # Try to find solved count - this at least confirms the user exists
            solved_text = soup.find(string=re.compile(r"Solved Problems"))
            if solved_text:
                parent = solved_text.parent
                count_elem = parent.find_next("div")
                if count_elem:
                    solved_count = count_elem.text.strip()
                    print(f"Found user profile. Solved count: {solved_count}")

                    # We found the profile but can't get individual problems
                    # Just return a marker that the user exists
                    return ["PROFILE_FOUND_BUT_NO_DETAILS"]

            # Try to extract from scripts
            scripts = soup.find_all("script")
            for script in scripts:
                if script.string and "solvedQuestions" in script.string:
                    match = re.search(r'"solvedQuestions":\s*(\[.*?\])', script.string)
                    if match:
                        try:
                            import json

                            solved_ids = json.loads(match.group(1))
                            for problem_id in solved_ids:
                                solved_problems.append({"id": str(problem_id)})

                            print(f"Found {len(solved_problems)} solved problem IDs")
                            return solved_problems
                        except json.JSONDecodeError:
                            pass

        # If we get here, all approaches failed
        print(
            "Could not retrieve solved problems. Please check if the username is correct."
        )
        return []

    except Exception as e:
        print(f"Error fetching solved problems: {e}")
        return []


def update_excel_with_status(
    excel_file, problem_titles, leetcode_problems, solved_problems
):
    """Update the Excel file with problem status, IDs, and difficulty"""
    print("Updating Excel file with problem status...")

    # Create a set of solved problem IDs for faster lookup
    solved_ids = {problem.get("id") for problem in solved_problems if "id" in problem}
    solved_titles = {
        problem.get("title") for problem in solved_problems if "title" in problem
    }

    # Check if we only found that the profile exists but not details
    profile_found = False
    if solved_problems and solved_problems[0] == "PROFILE_FOUND_BUT_NO_DETAILS":
        profile_found = True
        solved_ids = set()
        solved_titles = set()

    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Define color fills
        solved_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )  # Light green
        unsolved_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )  # Light yellow

        # Add headers for new columns if they don't exist
        if sheet.cell(row=1, column=2).value is None:
            sheet.cell(row=1, column=2).value = "Problem ID"
        if sheet.cell(row=1, column=3).value is None:
            sheet.cell(row=1, column=3).value = "Difficulty"
        if sheet.cell(row=1, column=4).value is None:
            sheet.cell(row=1, column=4).value = "Status"

        # Track how many problems were identified and marked
        identified_count = 0
        solved_count = 0
        total_problems = 0

        # Process each row (starting from row 2 to skip header)
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value

            if cell_value is None:
                continue

            total_problems += 1
            problem_id = None
            problem_title = None
            difficulty = None
            status = "Unknown"

            # Extract problem information from cell text
            cell_str = str(cell_value)

            # Try to extract problem ID from link if it exists
            id_match = re.search(r"problems/([^/]+)/", cell_str)

            if id_match:
                slug = id_match.group(1)

                # Look for problem in our map by slug or by checking title
                for title, info in leetcode_problems.items():
                    if not isinstance(info, dict):
                        continue

                    if "title" not in info and title.lower().replace(
                        "-", " "
                    ) == slug.lower().replace("-", " "):
                        problem_id = info.get("id")
                        difficulty = info.get("difficulty")
                        problem_title = title
                        break

            # If no ID found, try to match by title
            if not problem_id:
                # Try to extract title
                title_match = re.search(r">([^<]+)<", cell_str) or re.search(
                    r"([\w\s\(\)-]+)", cell_str
                )

                if title_match:
                    extracted_title = title_match.group(1).strip()

                    # Look for exact title match
                    if extracted_title in leetcode_problems:
                        info = leetcode_problems[extracted_title]
                        problem_id = info.get("id")
                        difficulty = info.get("difficulty")
                        problem_title = extracted_title
                    else:
                        # Try case-insensitive matching
                        for title, info in leetcode_problems.items():
                            if not isinstance(info, dict) or "title" in info:
                                continue

                            if title.lower() == extracted_title.lower():
                                problem_id = info.get("id")
                                difficulty = info.get("difficulty")
                                problem_title = title
                                break

            # If we identified the problem
            if problem_id or problem_title:
                identified_count += 1

                # Update problem ID and difficulty
                if problem_id:
                    sheet.cell(row=row, column=2).value = problem_id
                if difficulty:
                    sheet.cell(row=row, column=3).value = difficulty

                # Check if it's solved
                if problem_id in solved_ids or problem_title in solved_titles:
                    status = "Solved"
                    solved_count += 1
                    sheet.cell(row=row, column=4).value = status
                    # Apply green background
                    sheet.cell(row=row, column=1).fill = solved_fill
                    sheet.cell(row=row, column=4).fill = solved_fill
                elif (
                    not profile_found
                ):  # Only mark as unsolved if we could access solved problems
                    status = "Unsolved"
                    sheet.cell(row=row, column=4).value = status
                    # Apply yellow background
                    sheet.cell(row=row, column=1).fill = unsolved_fill
                    sheet.cell(row=row, column=4).fill = unsolved_fill

        # Save the updated workbook
        output_file = excel_file.replace(".xlsx", "_updated.xlsx")
        workbook.save(output_file)

        print(f"\nExcel file updated successfully!")
        print(f"Total problems in your list: {total_problems}")
        print(f"Problems identified in LeetCode database: {identified_count}")

        if profile_found:
            print(
                f"Found your LeetCode profile, but couldn't get detailed solved problem list."
            )
            print(f"You'll need to manually mark which problems you've solved.")
        else:
            print(f"Problems marked as solved: {solved_count}")

        print(f"Updated file saved as: {output_file}")

        return output_file

    except Exception as e:
        print(f"Error updating Excel file: {e}")
        print(f"Specific error: {type(e).__name__}: {str(e)}")
        return None


def main():
    print("\n=== LeetCode Progress Tracker ===\n")

    # Get Excel file path
    excel_file = input(
        "Enter the path to your Excel file (e.g., leetcode_problems.xlsx): "
    )

    if not os.path.exists(excel_file):
        print(f"File not found: {excel_file}")
        return

    # Get username
    username = input("Enter your LeetCode username: ")

    # Get the mapping of LeetCode problems
    leetcode_problems = get_leetcode_problems()

    if not leetcode_problems:
        print("Failed to fetch LeetCode problems. Cannot update tracking file.")
        return

    # Extract problem titles from Excel
    problem_titles = extract_problem_titles_from_excel(excel_file)

    if not problem_titles:
        print("Could not extract problem titles from your Excel file.")
        return

    # Get solved problems for user
    solved_problems = get_user_solved_problems(username)

    # Update Excel with status
    updated_file = update_excel_with_status(
        excel_file, problem_titles, leetcode_problems, solved_problems
    )

    if updated_file:
        print(
            f"\nYou can now open {updated_file} to see your updated LeetCode progress!"
        )


if __name__ == "__main__":
    main()
